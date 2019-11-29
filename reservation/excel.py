from openpyxl import Workbook, load_workbook, cell

file_path = r'data\test.xlsx'
wb = load_workbook(file_path)
if wb is None:
    print(f'Error opening the file: {file_path}')
# ws = wb.active
sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]

row , col=1,1
while True:
    mycell = ws.cell(row, col)
    field_value = mycell.value
    field_format = mycell.number_format
    print (f'{row}- value: {field_value}, format:{field_format}, type: {ws.cell(row, col).data_type}, number_format: {field_format}')
    row += 1
    if ws.cell(row,col).value is None:
        break

# field_value = ws.cell(2, col).value
# field_format = ws.cell(2, col).number_format
# print (field_value, field_format, ws.cell(2, col).data_type )
#
# field_value = ws.cell(3, col).value
# field_format = ws.cell(3, col).number_format
# print (field_value, field_format, ws.cell(3, col).data_type )

# field_type = str(type(field_value))
# mycell = ws.cell(row, col)
# field_value = mycell.value
# # check if the field is not string
# if mycell.data_type == cell.Cell.TYPE_NUMERIC:
