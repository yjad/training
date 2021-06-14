from re import findall
from openpyxl import Workbook, load_workbook
import tkinter
from tkinter import filedialog


def get_file_path(initialDir=".", DialogTitle="select File", fileTypeText="all Files", fileType="*.*"):
    root = tkinter.Tk()
    root.withdraw()
    file_path: str

    file_path = filedialog.askopenfilename(initialdir=initialDir, title=DialogTitle,
                                           filetypes=((fileTypeText, fileType), ("all files", "*.*")))

    return file_path


class ValidateUploadSheet:
    ws = None
    wb = None
    max_row = 65000
    header = []
    fp = None

    def __init__(self, file_path):
        self.wb = load_workbook(file_path)
        if self.wb is None:
            self.disp_general_error(f'Error opening the file: {file_path}')
        # ws = wb.active
        sheet_name = self.wb.sheetnames[0]
        self.ws = self.wb[sheet_name]

        if self.ws is None:
            self.disp_general_error(f'Error accessing sheet: {sheet_name}')

    def valid_str(self, in_str: str):
        if in_str.count('\n') > 0:
            return -1  # not valid, includes /r/n
        elif in_str.count(',') > 0:
            return -2  # not valid, includes comma
        else:
            return 0

    def disp_general_error(self, error_msg):
        print(f'Error: {error_msg}')
        self.fp.write(f'Error: {error_msg}\n')

    def disp_info(self, msg):
        print('Info: ' + msg)
        self.fp.write(f'Error: {msg}\n')

    #def disp_line_error(self, cell_row, field_name, field_value, opt_text=None):
    def disp_line_error(self, cell_row, cell_col, opt_text=None):
        field_name = self.header[cell_col-1]        # col in excel starts from 1
        field_value = self.ws.cell(row=cell_row,column=cell_col).value
        if opt_text is None:
            error_msg = f'Row {cell_row}: Invalid field: {field_name}: {field_value}'
        else:
            error_msg = f'Row {cell_row}: Invalid field: {field_name}: {field_value} --> {opt_text}'
        self.disp_general_error(error_msg)

    def validate_int_field(self, row, col):
        field_value = self.ws.cell(row, col).value
        if field_value > 255:
            self.disp_line_error(row, col, 'Int field cannot be more than 255 char')
            return False
        else:
            return True

    # area is big integer but should not include thousand seperator
    def validate_bigint_field(self, row, col):
        field_value = self.ws.cell(row, col).value
        field_format = self.ws.cell(row, col).number_format
        field_type = str(type(field_value))
        if 'float' in field_type:
            self.disp_line_error(row, col, 'field must be integer')
        elif 'int' in field_type:
            if field_value >= 1000 and field_format is not None:
                self.disp_line_error(row,  col, 'number & if formatted may contain comma')
            else:
                pass
        else :  # field is string
            #print ('field_type: ', field_type)
            self.validate_str_field(row, col)

    def validate_str_field(self, row, col):
        field_value = self.ws.cell(row, col).value
        # check if the field is not string
        if type(field_value) == int:
            field_value = str(field_value)
        if self.valid_str(field_value):
            self.disp_line_error(row, col, 'field includes newLine or comma')
            return False
        else:
            return True

    def validate_float(self, row, col, num_of_decimals):
        field_value = self.ws.cell(row, col).value
        if 'float' not in str(type(field_value)):
            self.disp_line_error(row, col, 'field should be decimal')
        # check number of decimals:
        x = str(field_value)
        if x[::-1].find('.') > num_of_decimals:
            self.disp_line_error(row, col, f'field has more than {num_of_decimals} decimal places')
            return False
        else:
            return True


class ValidateLandSheet(ValidateUploadSheet):
    sheet_no_of_columns = 13
    header = [
        'كود المحافظة', 'كود المدينة', 'المدينة', 'كود المنطقة', 'المنطقة', 'كود الحي', 'الحي', 'كود المجاورة',
        'المجاورة', 'رقم القطعة', 'مساحة القطعة', 'نسبة التميز', 'عدد القطع المتاحة'
    ]

    def check_file_header(self):

        for i in range (1, self.sheet_no_of_columns):
            if self.ws.cell(row=1, column=i).value != self.header[i-1]:
                self.disp_general_error (f'Invalid Header, column: {self.ws.cell.value}' )

    def check_rows(self):
        cell_row = 2
        while True:
            if self.ws.cell(row=cell_row, column=1).value is None:
                self.disp_info (f'No of lands : {cell_row-2}' )
                return

        # 'كود المحافظة'
            self.validate_int_field(cell_row, col=1)
        # city code
            self.validate_int_field(cell_row, col=2)
        # city
            self.validate_str_field(cell_row, col=3)
        # area Code
            self.validate_int_field(cell_row, col=4)
        # 'المنطقة'
            self.validate_str_field(cell_row, col=5)
        # check الحى
            self.validate_int_field(cell_row, col=6)
        # District name
            self.validate_str_field(cell_row, col=7)
        # sub district
            self.validate_int_field(cell_row, col=8)
        # sub District name
            self.validate_str_field(cell_row, col=9)
        # 'رقم القطعة', 'مساحة القطعة', 'نسبة التميز', 'عدد القطع المتاحة'
        # land-no is string, if number, make sure has no comma
            self.validate_bigint_field(cell_row, 9)
        # area is big number but should not include thousand seperator
            self.validate_bigint_field(cell_row, 10)
        # 'نسبة التميز'is decimal with only 2 decimal places
            self.validate_float(cell_row, 12, 2)
        # no of available lands should be 1
            self.validate_int_field(cell_row, col=13)

            if self.ws.cell(row=cell_row, column=13).value != 1:
                self.disp_line_error(cell_row, 13, 'no of lands should be 1')

            if self.ws.cell(row=cell_row, column=self.sheet_no_of_columns+1) is not None:
                self.disp_general_error(f'Row {cell_row} sheet should be {self.sheet_no_of_columns} column, there are other columns in the sheet')

            cell_row += 1
            if cell_row > self.max_row:
                return

    def check_land_sheet(self):
        self.check_file_header()
        self.check_rows()
        self.wb.close()
        self.fp.close()

def vlaidate_land_sheet():
    try:
        land_file_path=get_file_path(initialDir=".", DialogTitle="select excel sheet", fileTypeText="excel", fileType="*.xlsx")
    except:
        print ('No file selected ...')
    #land_file_path = r'E:/yahia/Python/training/reservation/مميزة 2 uploud.xlsx'
    else:
        if land_file_path == '':
            return
        validate_sheet = ValidateLandSheet(land_file_path)
        validate_sheet.fp = open(land_file_path+'.txt','w+', encoding='utf-8')
        validate_sheet.check_land_sheet()


vlaidate_land_sheet()