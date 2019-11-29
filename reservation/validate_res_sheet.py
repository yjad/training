from re import findall
from openpyxl import Workbook, load_workbook, cell
from tkinter import Tk, Label, Button, Entry, IntVar, END, W, E, NE, StringVar, DoubleVar, WORD
from tkinter import filedialog
from tkinter.scrolledtext import ScrolledText
import csv



def get_file_path(initialDir=".", DialogTitle="select File", fileTypeText="all Files", fileType="*.*"):
    #root = Tk()
    #root.withdraw()
    file_path: str

    file_path = filedialog.askopenfilename(initialdir=initialDir, title=DialogTitle,
                                           filetypes=((fileTypeText, fileType), ("all files", "*.*")))

    return file_path


class ValidateUploadSheet:
    ws = None
    wb = None
    max_row = 65000
    header = []
    fp = None
    output_win = None

    def __init__(self, file_path):
        self.wb = load_workbook(file_path)
        if self.wb is None:
            self.disp_general_error(f'Error opening the file: {file_path}')
        # ws = wb.active
        sheet_name = self.wb.sheetnames[0]
        self.ws = self.wb[sheet_name]

        if self.ws is None:
            self.disp_general_error(f'Error accessing sheet: {sheet_name}')

        try:
            self.fp = open(file_path + '.txt', 'w+', encoding='utf-8')
        except:
            self.disp_general_error(f'Error creating o/p file: {file_path}')


    def valid_str(self, in_str: str):
        if in_str.count('\n') > 0:
            return -1  # not valid, includes /r/n
        elif in_str.count(',') > 0:
            return -2  # not valid, includes comma
        else:
            return 0

    def disp_general_error(self, error_msg):
        msg  = f'Error: {error_msg}'
        #print(msg)
        self.fp.write(msg)
        self.output_win.insert(END, msg+'\n')

    def disp_info(self, in_msg):
        msg = 'Info: ' + in_msg
        #print(msg)
        self.fp.write(f'Error: {msg}\n')
        self.output_win.insert(END, msg+'\n')

    #def disp_line_error(self, cell_row, field_name, field_value, opt_text=None):
    def disp_line_error(self, cell_row, cell_col, opt_text=None):
        field_name = self.header[cell_col-1]        # col in excel starts from 1
        field_value = self.ws.cell(row=cell_row,column=cell_col).value
        if opt_text is None:
            error_msg = f'Row {cell_row}: Invalid field: {field_name}: {field_value}'
        else:
            error_msg = f'Row {cell_row}: Invalid field: {field_name}: {field_value} --> {opt_text}'
        self.disp_general_error(error_msg)

    def validate_small_int_field(self, row, col):
        field_value = self.ws.cell(row, col).value
        if field_value > 255:
            self.disp_line_error(row, col, 'Int field cannot be more than 255 char')
            return False
        else:
            return True

    # area is big integer but should not include thousand seperator
    def validate_int_field(self, row, col):
        field_value = self.ws.cell(row, col).value
        field_format = self.ws.cell(row, col).number_format
        field_type = str(type(field_value))
        if 'float' in field_type:
            self.disp_line_error(row, col, 'field must be integer')
        elif 'int' in field_type:
            if field_value >= 1000 and field_format != 'General':
                self.disp_line_error(row,  col, 'number & if formatted may contain comma')
            else:
                pass
        else :  # field is string
            #print ('field_type: ', field_type)
            self.validate_str_field(row, col)

    def validate_str_field(self, row, col):
        mycell = self.ws.cell(row, col)
        field_value = mycell.value
        # check if the field is not string
        if mycell.data_type == cell.cell.TYPE_NUMERIC:
            self.validate_int_field(row,col)
        elif self.valid_str(field_value):
            self.disp_line_error(row, col, 'field includes newLine or comma')
            return False
        else:
            return True

    def validate_float(self, row, col, num_of_decimals):
        field_value = self.ws.cell(row, col).value
        if 'float' not in str(type(field_value)):
            self.disp_line_error(row, col, 'field should be decimal')
        # check number of decimals:
        x = str(field_value)
        if x[::-1].find('.') > num_of_decimals:
            self.disp_line_error(row, col, f'field has more than {num_of_decimals} decimal places')
            return False
        else:
            return True
class ValidateUploadCSV:
    ws = None
    wb = None
    max_row = 65000
    header = []
    fp = None
    output_win = None

    def __init__(self, csv_file_path):
        csv_file = open(csv_file_path, mode='r', newline="\n", encoding='utf-8')
        csv_reader = csv.reader(csv_file, delimiter='|', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        # read header
        #csv_writer.writerow(['Quote', 'author', 'tag-1', 'tag-2', 'tag-3', 'tag-4', 'tag-5'])

        # self.wb = load_workbook(file_path)
        # if self.wb is None:
        #     self.disp_general_error(f'Error opening the file: {file_path}')
        # ws = wb.active
        # sheet_name = self.wb.sheetnames[0]
        # self.ws = self.wb[sheet_name]

        # if self.ws is None:
        #     self.disp_general_error(f'Error accessing sheet: {sheet_name}')

        try:
            self.fp = open(csv_file_path + '.txt', 'w+', encoding='utf-8')
        except:
            self.disp_general_error(f'Error creating o/p file: {csv_file_path}')


    def valid_str(self, in_str: str):
        if in_str.count('\n') > 0:
            return -1  # not valid, includes /r/n
        elif in_str.count(',') > 0:
            return -2  # not valid, includes comma
        else:
            return 0

    def disp_general_error(self, error_msg):
        msg  = f'Error: {error_msg}'
        #print(msg)
        self.fp.write(msg)
        self.output_win.insert(END, msg+'\n')

    def disp_info(self, in_msg):
        msg = 'Info: ' + in_msg
        #print(msg)
        self.fp.write(f'Error: {msg}\n')
        self.output_win.insert(END, msg+'\n')

    #def disp_line_error(self, cell_row, field_name, field_value, opt_text=None):
    def disp_line_error(self, cell_row, cell_col, opt_text=None):
        field_name = self.header[cell_col-1]        # col in excel starts from 1
        field_value = self.ws.cell(row=cell_row,column=cell_col).value
        if opt_text is None:
            error_msg = f'Row {cell_row}: Invalid field: {field_name}: {field_value}'
        else:
            error_msg = f'Row {cell_row}: Invalid field: {field_name}: {field_value} --> {opt_text}'
        self.disp_general_error(error_msg)

    def validate_small_int_field(self, row, col):
        field_value = self.ws.cell(row, col).value
        if field_value > 255:
            self.disp_line_error(row, col, 'Int field cannot be more than 255 char')
            return False
        else:
            return True

    # area is big integer but should not include thousand seperator
    def validate_int_field(self, row, col):
        field_value = self.ws.cell(row, col).value
        field_format = self.ws.cell(row, col).number_format
        field_type = str(type(field_value))
        if 'float' in field_type:
            self.disp_line_error(row, col, 'field must be integer')
        elif 'int' in field_type:
            if field_value >= 1000 and field_format != 'General':
                self.disp_line_error(row,  col, 'number & if formatted may contain comma')
            else:
                pass
        else :  # field is string
            #print ('field_type: ', field_type)
            self.validate_str_field(row, col)

    def validate_str_field(self, row, col):
        mycell = self.ws.cell(row, col)
        field_value = mycell.value
        # check if the field is not string
        if mycell.data_type == cell.cell.TYPE_NUMERIC:
            self.validate_int_field(row,col)
        elif self.valid_str(field_value):
            self.disp_line_error(row, col, 'field includes newLine or comma')
            return False
        else:
            return True

    def validate_float(self, row, col, num_of_decimals):
        field_value = self.ws.cell(row, col).value
        if 'float' not in str(type(field_value)):
            self.disp_line_error(row, col, 'field should be decimal')
        # check number of decimals:
        x = str(field_value)
        if x[::-1].find('.') > num_of_decimals:
            self.disp_line_error(row, col, f'field has more than {num_of_decimals} decimal places')
            return False
        else:
            return True


class ValidateLandSheet(ValidateUploadSheet):
    sheet_no_of_columns = 13
    header = [
        'كود المحافظة', 'كود المدينة', 'المدينة', 'كود المنطقة', 'المنطقة', 'كود الحي', 'الحي', 'كود المجاورة',
        'المجاورة', 'رقم القطعة', 'مساحة القطعة', 'نسبة التميز', 'عدد القطع المتاحة'
    ]

    def check_file_header(self):
        for i in range (1, self.sheet_no_of_columns):
            if self.ws.cell(row=1, column=i).value != self.header[i-1]:
                self.disp_general_error (f'Invalid Header, column: {self.ws.cell.value}' )

    def check_rows(self):
        cell_row = 2
        while True:
            if self.ws.cell(row=cell_row, column=1).value is None:
                self.disp_info (f'No of lands : {cell_row-2}' )
                return

        # 'كود المحافظة'
            self.validate_small_int_field(cell_row, col=1)
        # city code
            self.validate_small_int_field(cell_row, col=2)
        # city
            self.validate_str_field(cell_row, col=3)
        # area Code
            self.validate_small_int_field(cell_row, col=4)
        # 'المنطقة'
            self.validate_str_field(cell_row, col=5)
        # check الحى
            self.validate_small_int_field(cell_row, col=6)
        # District name
            self.validate_str_field(cell_row, col=7)
        # sub district
            self.validate_small_int_field(cell_row, col=8)
        # sub District name
            self.validate_str_field(cell_row, col=9)
        # 'رقم القطعة', 'مساحة القطعة', 'نسبة التميز', 'عدد القطع المتاحة'
        # land-no is string, if number, make sure has no comma
            self.validate_int_field(cell_row, 9)
        # area is big number but should not include thousand seperator
            self.validate_int_field(cell_row, 10)
        # 'نسبة التميز'is decimal with only 2 decimal places
            self.validate_float(cell_row, 12, 2)
        # no of available lands should be 1
            self.validate_small_int_field(cell_row, col=13)

            if self.ws.cell(row=cell_row, column=13).value != 1:
                self.disp_line_error(cell_row, 13, 'no of lands should be 1')

            x= self.ws.cell(row=cell_row, column=self.sheet_no_of_columns+110)
            print (x.value, " date_type: ", x.data_type, x)

            #print (f'type: {type(x)}, format: {x.string}')
            if self.ws.cell(row=cell_row, column=self.sheet_no_of_columns+1).value != None:
                self.disp_general_error(f'Row {cell_row} sheet should be {self.sheet_no_of_columns} column, there are other columns in the sheet')

            cell_row += 1
            if cell_row > self.max_row:
                return

    def check_land_sheet(self):
        self.check_file_header()
        self.check_rows()
        self.wb.close()
        self.fp.close()


class MainMenu():

    def __init__(self, master):
        self.master = master
        self.output_file_name = StringVar()

        master.title ('Validate Reservation Sheet')
        #print(master.configure().keys())

        #Label(master, text='\nOutput',  fg='black', font='none 12 bold').grid(row=0, column=0, sticky = W)

        #self.output_file_name. = 'Output'
        Label(master, textvariable=self.output_file_name,  fg='black', font='none 12 bold').grid(row=0, column=0, sticky = W)

        self.output = ScrolledText(master, width=120, height = 40, wrap = WORD, background = 'white')
        self.output.grid(row=1, column = 0, columnspan = 2, sticky = 'W')

        land_btn = Button(master, text="Lands", command = self.validate_land_sheet).grid(row=0, column=2)
        unit_btn = Button(master, text="Units", command = self.val_unit).grid(row=0, column=3)

        exit_btn = Button(master, text="Exit ", command = master.quit).grid(row=0, column=4, sticky = W)

    def validate_land_sheet(self):
        self.output.delete(0.0, END)        # clear window
        file_path = self.select_file_path()
        self.output_file_name.set('Results of file name: '+ file_path)
        if file_path != '':
            val_land = ValidateLandSheet(file_path)
            val_land.output_win = self.output
            val_land.check_land_sheet()


    def val_unit(self):
        pass

    def select_file_path(self):
        try:
            file_path = get_file_path(initialDir=".", DialogTitle="select excel sheet", fileTypeText="excel",
                                           fileType="*.xlsx")
        except:
            print('No file selected ...')
            return ''
        else:
            return file_path

def validate_res_sheet():
    root = Tk()
    #root.attributes('-fullscreen', True)
    root.state('zoomed')
    MainMenu(root)
    root.mainloop()

validate_res_sheet()