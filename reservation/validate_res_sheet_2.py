from re import findall
from openpyxl import Workbook, load_workbook, cell
from tkinter import Tk, Label, Button, Entry, IntVar, END, W, E, NE, StringVar, DoubleVar, WORD
from tkinter import filedialog
from tkinter.scrolledtext import ScrolledText
import csv

EXCEL = 'xlsx'
CSV = 'csv'
NUMERIC_TYPE = 'n'
STRING_TYPE= 's'

def get_file_path(initialDir=".", DialogTitle="select File", fileTypeText="all Files", fileType="*.*"):
    #root = Tk()
    #root.withdraw()
    file_path: str

    file_path = filedialog.askopenfilename(initialdir=initialDir, title=DialogTitle,
                                           filetypes=((fileTypeText, fileType), ("all files", "*.*")))

    return file_path


class UploadSheetUtils:
    ws = None
    wb = None
    csv_file = None
    csv_reader = None
    max_row = 65000
    header = []
    fp = None
    output_win = None
    file_type = None
    curr_row_no = 0
    curr_row=[]

    def __init__(self, file_path):
        if file_path.split('.')[-1] == 'xlsx'.lower():
            self.file_type = EXCEL
            self.wb = load_workbook(file_path)
            if self.wb is None:
                self.disp_general_error(f'Error opening the file: {file_path}')
            # ws = wb.active
            sheet_name = self.wb.sheetnames[0]
            self.ws = self.wb[sheet_name]
            if self.ws is None:
                self.disp_general_error(f'Error accessing sheet: {sheet_name}')
        else:
            self.file_type = CSV
            try:
                self.csv_file = open(file_path, mode='r', newline="\n", encoding='utf-8')
            except:
                self.disp_general_error(f'Error opening file: {file_path}')

            self.ws = csv.reader(self.csv_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            if self.ws is None:
                self.disp_general_error(f'reading csv file: {file_path}')

        # Open O/p file
        try:
            self.fp = open(file_path + '.txt', 'w+', encoding='utf-8')
        except:
            self.disp_general_error(f'Error creating o/p file: {file_path}')

    def cell_value(self, row, col):
        if self.file_type == CSV:
            if row != self.curr_row_no:
                self.curr_row_no = row
                try:
                    self.curr_row = next(self.ws)
                except StopIteration: # end
                    return None
            #print (f'row: {row}, col: {col}, val: {self.curr_row[col - 1]}')
            return self.curr_row[col - 1].strip()
        else:   # excel
            return self.ws.cell(row=row, column= col).value

    def cell_type(self, row, col):
        if self.file_type == CSV:
            try:
                f:float = float()
                return NUMERIC_TYPE
            except:
                return STRING_TYPE
        else:
            return self.ws.cell(row=row, column= col).data_type

    def cell_format(self, row, col):
        if self.file_type == CSV:
            return None
        else:
            if self.ws.cell(row,col).number_format == 'General':
                return None
            else:
                return self.ws.cell(row=row, column= col).number_format

    def disp_general_error(self, error_msg):
        msg  = f'Error: {error_msg}'
        #print(msg)
        self.fp.write(msg)
        self.output_win.insert(END, msg+'\n')

    def disp_info(self, in_msg):
        msg = 'Info: ' + in_msg
        #print(msg)
        self.fp.write(f'Error: {msg}\n')
        self.output_win.insert(END, msg+'\n')

    def disp_line_error(self, cell_row, cell_col, opt_text=None):
        field_name = self.header[cell_col-1]        # col in excel starts from 1
        field_value = self.cell_value (cell_row, cell_col)
        if opt_text is None:
            error_msg = f'Row {cell_row}: Invalid field: {field_name}: {field_value}'
        else:
            error_msg = f'Row {cell_row}: Invalid field: {field_name}: {field_value} --> {opt_text}'
        self.disp_general_error(error_msg)

    # def validate_small_int_field(self, row, col):
    #     field_value = self.cell_value(row, col)
    #     if field_value > 255:
    #         self.disp_line_error(row, col, 'Int field cannot be more than 255')
    #         return False
    #     else:
    #         return True

    # area is big integer but should not include thousand seperator
    def validate_numeric_field(self, row, col, max_number=None):
        f = self.cell_value(row, col)
        field_value = self.cell_value(row, col)
        field_format = self.cell_format(row, col)
        field_type = self.cell_type(row, col)
        print (field_type)
        if field_type == STRING_TYPE:
            self.disp_line_error(row, col, 'field must be integer')
            return False
        elif field_type != NUMERIC_TYPE:
            self.disp_line_error(row, col, 'field must be integer')
            return False
        if max_number is not None:
            if int(field_value) > max_number:
                self.disp_line_error(row, col, f'Int field cannot be more than {max_number}')
                return False
            elif int(field_value) >= 1000 and field_format is not None:
                self.disp_line_error(row,  col, 'number & if formatted may contain comma')
                return False
            else:
                return True

    def validate_str_field(self, row, col):

        field_value = self.cell_value(row, col)
        field_type = self.cell_type(row, col)

        if field_type == NUMERIC_TYPE:
            return self.validate_numeric_field(row, col, None)
        else:
            # check if the field is not string
            if field_value.count('\n') > 0:
                self.disp_line_error(row, col, 'field includes newLine')
            elif field_value.count(',') > 0:
                self.disp_line_error(row, col, 'field includes comma')
            else:
                return True

    def validate_float(self, row, col, num_of_decimals):
        field_value = self.cell_value(row, col)
        if self.cell_type(row,col) == STRING_TYPE:
            self.disp_line_error(row, col, 'field should be decimal')
        # check number of decimals:
        x = str(field_value)
        if x[::-1].find('.') > num_of_decimals:
            self.disp_line_error(row, col, f'field has more than {num_of_decimals} decimal places')
            return False
        else:
            return True

    def close_files(self):
        if self.file_type == CSV:
            self.csv_file.close()
        else:
            self.wb.close()
        self.fp.close()


class LandSheetUpload(UploadSheetUtils):
    sheet_no_of_columns = 13
    header = [
        'كود المحافظة', 'كود المدينة', 'المدينة', 'كود المنطقة', 'المنطقة', 'كود الحي', 'الحي', 'كود المجاورة',
        'المجاورة', 'رقم القطعة', 'مساحة القطعة', 'نسبة التميز', 'عدد القطع المتاحة'
    ]

    def check_file_header(self):
        for i in range(1, self.sheet_no_of_columns):
            if self.cell_value(row=1, col=i) != self.header[i - 1]:
                self.disp_general_error(f'Invalid Header, column: {self.cell_value(1,i)}')

    def check_rows(self):
        cell_row = 2
        while True:
            #print ('cell row:', cell_row, self.curr_row_no)
            if self.cell_value(row=cell_row, col=1) is None:
                self.disp_info(f'No of lands : {cell_row - 2}')
                return

            # 'كود المحافظة'
            self.validate_numeric_field(cell_row, col=1, max_number=255)
            # city code
            self.validate_numeric_field(cell_row, col=2, max_number=255)
            # city
            self.validate_str_field(cell_row, col=3)
            # area Code
            self.validate_numeric_field(cell_row, col=4, max_number=255)
            # 'المنطقة'
            self.validate_str_field(cell_row, col=5)
            # check الحى
            self.validate_numeric_field(cell_row, col=6, max_number=255)
            # District name
            self.validate_str_field(cell_row, col=7)
            # sub district
            self.validate_numeric_field(cell_row, col=8, max_number=255)
            # sub District name
            self.validate_str_field(cell_row, col=9)
            # 'رقم القطعة', 'مساحة القطعة', 'نسبة التميز', 'عدد القطع المتاحة'
            # land-no is string, if number, make sure has no comma
            self.validate_numeric_field(cell_row, 10)
            # area is big number but should not include thousand seperator
            self.validate_numeric_field(cell_row, 11)
            # 'نسبة التميز'is decimal with only 2 decimal places
            self.validate_float(cell_row, 12, 2)
            # no of available lands should be 1
            self.validate_numeric_field(cell_row, col=13, max_number=1)

            # if self.ws.cell(row=cell_row, column=self.sheet_no_of_columns+1).value != "":
            #     self.disp_general_error(f'Row {cell_row} sheet should be {self.sheet_no_of_columns} column, there are other columns in the sheet')

            if self.file_type == CSV:
                if len(self.curr_row) > self.sheet_no_of_columns:
                    self.disp_general_error(
                        f'Row {cell_row} sheet should be {self.sheet_no_of_columns} column, there are other columns in the sheet')
                elif len(self.curr_row) < self.sheet_no_of_columns:
                    self.disp_general_error(
                        f'Row {cell_row} sheet should be {self.sheet_no_of_columns} column, there are less columns in the sheet')
            cell_row += 1
            if cell_row > self.max_row:
                return

    def check_land_sheet(self):
        self.check_file_header()
        self.check_rows()
        self.close_files()


class MainMenu():

    def __init__(self, master):
        self.master = master
        self.output_file_name = StringVar()

        master.title ('Validate Reservation Sheet')
        #print(master.configure().keys())

        #Label(master, text='\nOutput',  fg='black', font='none 12 bold').grid(row=0, column=0, sticky = W)

        #self.output_file_name. = 'Output'
        Label(master, textvariable=self.output_file_name,  fg='black', font='none 12 bold').grid(row=0, column=0, sticky = W)

        self.output = ScrolledText(master, width=120, height = 40, wrap = WORD, background = 'white')
        self.output.grid(row=1, column = 0, columnspan = 2, sticky = 'W')

        land_btn = Button(master, text="Lands", command = self.validate_land_sheet).grid(row=0, column=2)
        unit_btn = Button(master, text="Units", command = self.val_unit).grid(row=0, column=3)

        exit_btn = Button(master, text="Exit ", command = master.quit).grid(row=0, column=4, sticky = W)

    def validate_land_sheet(self):
        file_path = self.select_file_path()
        if file_path != '':
            val_land = LandSheetUpload(file_path)

            val_land.output_win = self.output
            val_land.output_win.delete(0.0, END)        # clear window
            self.output_file_name.set('Results of file name: '+ file_path)
            val_land.check_land_sheet()

    def val_unit(self):
        pass

    def select_file_path(self):
        try:
            file_path = get_file_path(initialDir=".", DialogTitle="select land sheet", fileTypeText="excel or csv",
                                           fileType="*.xlsx;*.csv")
        except:
            print('No file selected ...')
            return ''
        else:
            return file_path

def validate_res_sheet():
    root = Tk()
    #root.attributes('-fullscreen', True)
    root.state('zoomed')
    MainMenu(root)
    root.mainloop()

validate_res_sheet()