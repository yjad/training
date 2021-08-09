import sqlite3
import sys
import os 
from openpyxl import Workbook
from tkinter import END, messagebox
from config import CONFIG
from tabulate import tabulate

DATA_FOLDER = CONFIG.get("DATA_FOLDER")
OUT_FOLDER = CONFIG.get("OUT_FOLDER")
DB_FILE_NAME = CONFIG.get("DB_FILE_NAME")
OUTPUT = None

def open_db():
    if not os.path.exists(DB_FILE_NAME):
        print_tk (f"database file {DB_FILE_NAME} does not exist")
        connection, cursor = create_db()
    else:    
        connection = sqlite3.Connection(DB_FILE_NAME)
        cursor = connection.cursor()
    return connection, cursor



def create_db():
    connection = sqlite3.Connection(DB_FILE_NAME)
    cursor = connection.cursor()
    
    with open (os.path.join(DATA_FOLDER, "HR.sqlite.sql")) as f:
        sql = f.read()
        cursor.executescript(sql)
        # conn.commit()
        # close_db(cursor)
        return connection, cursor 
    
def close_db(cursor):
    cursor.close()


def insert_row_dict(conn, cursor, table_name, rec):
    keys = ','.join(rec.keys())
    question_marks = ','.join(list('?' * len(rec)))
    values = tuple(rec.values())
    try:
        cursor.execute('INSERT INTO ' + table_name + ' (' + keys + ') VALUES (' + question_marks + ')', values)
        conn.commit()
        return 0
    except sqlite3.Error as er:
        print('SQLite error: %s' % (' '.join(er.args)))
        print("Exception class is: ", er.__class__)
        print('SQLite traceback: ')
        exc_type, exc_value, exc_tb = sys.exc_info()
        print(exc_type, exc_value, exc_tb, "\n", "table: ", table_name, "--", rec)
        # print(traceback.format_exception(exc_type, exc_value, exc_tb))
        return -1

def insert_row_list(conn, cursor, table_name, rec):


    question_marks = ','.join(list('?' * len(rec)))
    try:
        cursor.execute('INSERT INTO ' + table_name + ' VALUES (' + question_marks +')', tuple(rec))
        # conn.commit()
        return 0
    except sqlite3.Error as er:
        print('SQLite error: %s' % (' '.join(er.args)))
        print("Exception class is: ", er.__class__)
        print('SQLite traceback: ')
        exc_type, exc_value, exc_tb = sys.exc_info()
        print(traceback.format_exception(exc_type, exc_value, exc_tb))
        
        print(exc_type, exc_value, exc_tb)
        # print_tk(exc_type, exc_value, exc_tb)
        exit(-1)
        # return -1
        
        
def exec_db_cmd(cmd):
    
    conn, cursor = open_db()
    try:
        cursor.execute(cmd)
        conn.commit()
        close_db(cursor)
        return 0
    except sqlite3.Error as er:
        exc_type, exc_value, exc_tb = sys.exc_info()
        if str(exc_value) == 'database is locked':
            messagebox.showinfo("Error!",exc_value)
            return -1
        else:
            print('SQLite error: %s' % (' '.join(er.args)))
            print("Exception class is: ", er.__class__)
            print('SQLite traceback: ')
            print(exc_type, exc_value, exc_tb, "\n", cmd)
            exit(-1)
        

def exec_query(cursor, cmd):
    cursor.execute(cmd)
    rows = cursor.fetchall()
    return rows


def get_col_names(conn, sql):
    get_column_names = conn.execute(sql + " limit 1")
    col_name = [i[0] for i in get_column_names.description]
    return col_name
    

def query_to_excel(cmd, file_name, header=None):
    conn, cursor = open_db()
    rows = exec_query(cursor, cmd)

    if header:
        headers = get_col_names(conn, cmd)
    close_db(cursor)
    
    wb = Workbook()
    ws = wb.active
    ws.title = os.path.splitext(os.path.basename(file_name))[0]
    ws.append(headers)

    for row in rows:
        ws.append(row)
    try:
        wb.save(file_name)
    except:
        r = messagebox.askretrycancel ("error saving to excel file ...")
        print_tk ("return code: ", r)

    # messagebox.showinfo("Done!",f"file exported {file_name}! ")
    

def query_to_list(cmd, return_header = True):
    conn, cursor = open_db()
    rows = exec_query(cursor, cmd)
    if return_header:
        if cmd.upper().find("LIMIT") != -1:   # if command have "Limit clause, dont return header
            header = []
        else:
            header = get_col_names(conn, cmd)
        close_db(cursor)
        return header, rows
    else:
        close_db(cursor)
        return rows
        
def query_to_dict_list(cmd):
    conn = sqlite3.Connection(DB_FILE_NAME)
    conn.row_factory = sqlite3.Row      # fetch as dict
    cursor = conn.cursor()
    cursor.execute(cmd)
    rows = [dict(row) for row in cursor.fetchall()]
    
    close_db(cursor)
    return rows
        
def set_output_tk(out):
    global OUTPUT
    OUTPUT = out
    
def clear_tk():
    OUTPUT.delete(0.0, END)        # clear window
 
 
def print_tk(*args):
    s = ''
    for a in args:
        s = s + str(a) + " "
        
    OUTPUT.insert(END, f"{s}\n")
        

def wrap(string, lenght=15):
    return '\n'.join(textwrap.wrap(string, lenght))

    
def print_query(cmd, title, max_len=30, export=True, confirmation=True):
    
    MAX_DISPLAY_LINES = 100
    
    if type(cmd) == str:
        header, rows = query_to_list(cmd)
    else:
        header, rows = cmd()        # a function that returns header & rows
        # print ("return back from cmd,", len(rows), "\n", rows)

    if not export:
        if max_len:     # wrap long fields at 30 chars
            new_rows = []
            for i,r in enumerate(rows):
                w_row = []
                for j, c in enumerate(r):
                    if len (str(c)) > max_len:
                        w_row.append(wrap(c, max_len))
                    else:
                        w_row.append(c)
                new_rows.append(w_row)
                if i > MAX_DISPLAY_LINES:   # display max of 100 lines
                    break

            table = tabulate(new_rows, headers=header, showindex="always")
        else:
            table = tabulate(rows, headers=header, showindex="always")
            
        clear_tk()      # clear Window
        output.insert(END, title+ '\n\n')
        for line in table:
            output.insert(END, line)
    else:   # export to excel
        if not os.path.exists(OUT_FOLDER):
            messagebox.showinfo("Error", f"Folder does not exist: {OUT_FOLDER}")
            return -1

        wb = Workbook()
        ws = wb.active
        ws.title = title
        ws.append(header)

        for row in rows:
            ws.append(row)
        # ws['G2'].style = 'Percent'
        file_name = os.path.join(OUT_FOLDER, title+".xlsx")
        while True:
            try:
                wb.save(os.path.join(file_name))
                break
            except:
                r = messagebox.askretrycancel ("Error", "Error saving to excel file, file is locked ...")
                if r == False:   # Cancel
                    return

        if confirmation:
            messagebox.showinfo("Done!",f"file exported {file_name}! ")

# this functions strip all white spaces even inside the text 

def trim(s):

    x = ' '.join(s.split())
    return x