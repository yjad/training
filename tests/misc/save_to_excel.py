from openpyxl import Workbook


def save_list_to_excel(header, list_of_values, file_name:str):

    wb = Workbook()

    ws = wb.active
    # rename sheet title
    ws.title = file_name.split('.')[0]
    #print (wb.sheetnames)
    #ws = wb.create_sheet(file_name.split('.')[0])
    # write headers

    cell_address_row = 1
    cell_address_col = 1
    for h in header:
        ws.cell(row=cell_address_row, column=cell_address_col, value=h)
        cell_address_col += 1

    # write values from rows of 'list_of_values'
    cell_address_row = 2
    for row in list_of_values:
        cell_address_col = 1
        for col in row:
            ws.cell(row=cell_address_row, column=cell_address_col, value=col)
            cell_address_col += 1
        cell_address_row += 1

    wb.save(file_name)


# test

# h = ['a', 'b', 'c','d']
# values = [
#     [1,2,3,4],
#     [5,6,7,8],
#     [9,10,11,12]
# ]
# save_to_excel (h, values, 'yahia.xlsx')


