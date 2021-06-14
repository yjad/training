from openpyxl import Workbook, load_workbook

file_path = r'D:\Yahia\Yahia\Python\training\reservation\مميزة 2 uploud.xlsx'
#file_path = r'..\reservation\extra.xlsx'
wb = load_workbook(file_path)

sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]

#Cell.TYPE_STRING
c1 = ws.cell(row=3, column=1)
c50 =  ws.cell(row=3, column=50)
c14 =  ws.cell(row=3, column=14)
c15 =  ws.cell(row=3, column=15)

print (f'c1 value: {c1.value}, c1.data_type = {c1.data_type}')
for i in range (10, 70):
    v = ws.cell(row=3, column=i).value
    t = ws.cell(row=3, column=i).data_type
    print (f'c{i} value: {v}, c{i}.data_type = {t}, {ws.cell(row=3, column=i)._value} ' )
print (f'c50 value: {c50.value}, c50.data_type = {c50.data_type}, {c50._value}')
print (ws.max_row, " -", ws.max_column)