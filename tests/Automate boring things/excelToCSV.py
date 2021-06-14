from openpyxl import Workbook, load_workbook
#import csv

sep = ','
file_name = r'D:\Yahia\Yahia\Python\training\reservation\مميزة 2 uploud.xlsx'

wb = load_workbook(file_name)

def extract_one_sheet(sheet_name):
    ws = wb[sheet_name]
    csv_file = open(file_name+'.csv', 'w+')

    for r in range(1, ws.max_row+1):
        line = ''
        for c in range (1,ws.max_column+1):
            if ws.cell(row=r, column=c).value is not None:
                line = line + sep + str(ws.cell(row=r, column=c).value)
        csv_file.write(line+'\n')
    csv_file.close()


for sheet in wb.sheetnames:
    extract_one_sheet(sheet)

